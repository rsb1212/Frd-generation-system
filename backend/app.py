import os
import logging
from dotenv import load_dotenv

# Load environment variables FIRST
load_dotenv()

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
from datetime import datetime
import json

from config import get_config
from rag_agent import RAGAgent
from document_processor import DocumentProcessor
from openai import AzureOpenAI

# Initialize Azure OpenAI client AFTER loading env vars
client = AzureOpenAI(
    azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
    api_key=os.getenv("AZURE_OPENAI_API_KEY"),
    api_version=os.getenv("AZURE_OPENAI_API_VERSION")
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)
config = get_config()
app.config.from_object(config)

# Initialize CORS
CORS(app, resources={r"/api/*": {"origins": config.CORS_ORIGINS}})

# Initialize services
try:
    rag_agent = RAGAgent(config)
    doc_processor = DocumentProcessor(config)
    logger.info("Services initialized successfully")
except Exception as e:
    logger.error(f"Failed to initialize services: {str(e)}")
    raise

# Create necessary directories
os.makedirs(config.UPLOAD_FOLDER, exist_ok=True)
os.makedirs(config.GENERATED_FOLDER, exist_ok=True)

# ==================== Health Check ====================
@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'service': 'FRD Generation System'
    }), 200

# ==================== Requirement Processing ====================
@app.route('/api/requirements/analyze', methods=['POST'])
def analyze_requirement():
    """Analyze and understand business requirement"""
    try:
        data = request.get_json()
        requirement = data.get('requirement')
        
        if not requirement:
            return jsonify({'error': 'Requirement text is required'}), 400
        
        # Validate requirement
        is_valid, issues = rag_agent.validate_requirement(requirement)
        if not is_valid:
            return jsonify({'error': 'Invalid requirement', 'issues': issues}), 400
        
        # Understand requirement using GPT
        analysis = rag_agent.understand_requirement(requirement)
        
        return jsonify({
            'success': True,
            'analysis': analysis,
            'timestamp': datetime.now().isoformat()
        }), 200
    
    except Exception as e:
        logger.error(f"Error in analyze_requirement: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/requirements/retrieve-similar', methods=['POST'])
def retrieve_similar():
    """Retrieve similar documents from database"""
    try:
        data = request.get_json()
        query = data.get('query')
        n_results = data.get('n_results', 5)
        
        if not query:
            return jsonify({'error': 'Query is required'}), 400
        
        similar_docs = rag_agent.retrieve_similar_documents(query, n_results)
        
        return jsonify({
            'success': True,
            'similar_documents': similar_docs,
            'count': len(similar_docs)
        }), 200
    
    except Exception as e:
        logger.error(f"Error in retrieve_similar: {str(e)}")
        return jsonify({'error': str(e)}), 500

# ==================== Document Generation ====================
@app.route('/api/frd/generate', methods=['POST'])
def generate_frd():
    """Generate complete FRD document"""
    try:
        data = request.get_json()
        requirement = data.get('requirement')
        user_info = data.get('user_info', {})
        output_format = data.get('format', 'docx')  # docx, pdf, or both
        include_templates = data.get('include_templates', False)
        
        if not requirement:
            return jsonify({'error': 'Requirement is required'}), 400
        
        # Validate
        is_valid, issues = rag_agent.validate_requirement(requirement)
        if not is_valid:
            return jsonify({'error': 'Invalid requirement', 'issues': issues}), 400
        
        # Retrieve similar documents
        similar_docs = rag_agent.retrieve_similar_documents(requirement, n_results=3)
        
        # Generate FRD sections
        logger.info("Generating FRD sections...")
        frd_sections = rag_agent.generate_frd_sections(requirement, similar_docs)
        
        if 'error' in frd_sections:
            return jsonify({'error': 'Failed to generate FRD sections'}), 500
        
        # Generate DOCX
        logger.info("Generating DOCX document...")
        docx_path = doc_processor.generate_docx(requirement, frd_sections, user_info)
        
        result = {
            'success': True,
            'requirement': requirement,
            'sections': frd_sections,
            'files': {
                'docx': os.path.basename(docx_path)
            },
            'timestamp': datetime.now().isoformat()
        }
        
        # Generate PDF if requested
        if output_format in ['pdf', 'both']:
            try:
                logger.info("Generating PDF document...")
                pdf_path = doc_processor.generate_pdf(docx_path)
                result['files']['pdf'] = os.path.basename(pdf_path)
            except Exception as e:
                logger.warning(f"PDF generation failed: {str(e)}")
                result['warnings'] = ['PDF generation failed, DOCX available']
        
        # Store in database for future retrieval
        try:
            doc_id = datetime.now().strftime('%Y%m%d%H%M%S')
            rag_agent.add_document_to_db(doc_id, requirement, {'type': 'frd', 'format': 'requirement'})
        except Exception as e:
            logger.warning(f"Failed to store in database: {str(e)}")
        
        return jsonify(result), 200
    
    except Exception as e:
        logger.error(f"Error in generate_frd: {str(e)}")
        return jsonify({'error': str(e)}), 500

# ==================== File Management ====================
@app.route('/api/files/upload', methods=['POST'])
def upload_file():
    """Upload FRD or template file"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Check file extension
        if not '.' in file.filename:
            return jsonify({'error': 'Invalid file format'}), 400
        
        ext = file.filename.rsplit('.', 1)[1].lower()
        if ext not in config.ALLOWED_EXTENSIONS:
            return jsonify({
                'error': f'File type not allowed. Allowed: {", ".join(config.ALLOWED_EXTENSIONS)}'
            }), 400
        
        # Check file size
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        
        if file_size > config.MAX_FILE_SIZE:
            return jsonify({'error': 'File size exceeds maximum allowed'}), 400
        
        # Save file
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{timestamp}_{filename}"
        filepath = os.path.join(config.UPLOAD_FOLDER, filename)
        file.save(filepath)
        
        # Extract and store content
        try:
            content = _extract_file_content(filepath, ext)
            if content:
                rag_agent.add_document_to_db(filename, content, {'type': 'uploaded', 'extension': ext})
        except Exception as e:
            logger.warning(f"Could not index uploaded file: {str(e)}")
        
        return jsonify({
            'success': True,
            'filename': filename,
            'size': file_size,
            'timestamp': datetime.now().isoformat()
        }), 200
    
    except Exception as e:
        logger.error(f"Error in upload_file: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/files/download/<filename>', methods=['GET'])
def download_file(filename):
    """Download generated file"""
    try:
        filepath = os.path.join(config.GENERATED_FOLDER, secure_filename(filename))
        
        if not os.path.exists(filepath):
            return jsonify({'error': 'File not found'}), 404
        
        return send_file(filepath, as_attachment=True)
    
    except Exception as e:
        logger.error(f"Error in download_file: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/files/list', methods=['GET'])
def list_files():
    """List generated files"""
    try:
        files = []
        for filename in os.listdir(config.GENERATED_FOLDER):
            filepath = os.path.join(config.GENERATED_FOLDER, filename)
            if os.path.isfile(filepath):
                files.append({
                    'filename': filename,
                    'size': os.path.getsize(filepath),
                    'created': datetime.fromtimestamp(os.path.getctime(filepath)).isoformat()
                })
        
        return jsonify({'success': True, 'files': files}), 200
    
    except Exception as e:
        logger.error(f"Error in list_files: {str(e)}")
        return jsonify({'error': str(e)}), 500

# ==================== Template Management ====================
@app.route('/api/templates/list', methods=['GET'])
def list_templates():
    """List available templates"""
    try:
        templates = [
            {'name': 'Enterprise', 'description': 'Large scale enterprise systems'},
            {'name': 'SaaS', 'description': 'Software as a Service applications'},
            {'name': 'Mobile', 'description': 'Mobile applications'},
            {'name': 'Web', 'description': 'Web applications'},
        ]
        return jsonify({'success': True, 'templates': templates}), 200
    
    except Exception as e:
        logger.error(f"Error in list_templates: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/templates/get/<template_name>', methods=['GET'])
def get_template(template_name):
    """Get template details"""
    try:
        templates = {
            'Enterprise': {
                'sections': [
                    'Project Overview', 'Scope', 'Functional Requirements',
                    'Non-Functional Requirements', 'User Roles', 'Use Cases'
                ],
                'description': 'Comprehensive template for enterprise systems'
            },
            'SaaS': {
                'sections': ['Project Overview', 'Scope', 'Core Features', 'User Workflows', 'APIs'],
                'description': 'Streamlined template for SaaS applications'
            }
        }
        
        template = templates.get(template_name)
        if not template:
            return jsonify({'error': 'Template not found'}), 404
        
        return jsonify({'success': True, 'template': template}), 200
    
    except Exception as e:
        logger.error(f"Error in get_template: {str(e)}")
        return jsonify({'error': str(e)}), 500

# ==================== Utility Functions ====================
def _extract_file_content(filepath: str, ext: str) -> str:
    """Extract text content from uploaded file"""
    try:
        if ext == 'txt':
            with open(filepath, 'r', encoding='utf-8') as f:
                return f.read()
        elif ext == 'docx':
            from docx import Document
            doc = Document(filepath)
            return '\n'.join(p.text for p in doc.paragraphs)
        elif ext == 'pdf':
            try:
                from pypdf import PdfReader
                reader = PdfReader(filepath)
                return '\n'.join(page.extract_text() for page in reader.pages)
            except:
                logger.warning("pypdf not available for PDF extraction")
                return ""
        elif ext in ['xlsx', 'xls']:
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(filepath, data_only=True)
                content = []
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    content.append(f"=== Sheet: {sheet_name} ===")
                    for row in sheet.iter_rows(values_only=True):
                        row_text = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                        if row_text.strip():
                            content.append(row_text)
                return '\n'.join(content)
            except Exception as e:
                logger.warning(f"openpyxl not available for Excel extraction: {e}")
                return ""
    except Exception as e:
        logger.error(f"Error extracting content: {str(e)}")
        return ""

# ==================== Error Handlers ====================
@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Endpoint not found'}), 404

@app.errorhandler(500)
def internal_error(error):
    logger.error(f"Internal server error: {str(error)}")
    return jsonify({'error': 'Internal server error'}), 500

@app.errorhandler(403)
def forbidden(error):
    return jsonify({'error': 'Forbidden'}), 403

# ==================== CLI Runner ====================
if __name__ == '__main__':
    env = os.getenv('FLASK_ENV', 'development')
    debug = env == 'development'
    port = int(os.getenv('PORT', 5000))
    
    logger.info(f"Starting Flask app in {env} mode on port {port}")
    app.run(host='0.0.0.0', port=port, debug=debug)
